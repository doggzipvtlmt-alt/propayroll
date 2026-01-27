from io import BytesIO
from typing import Iterable

from openpyxl import Workbook, load_workbook
from django.http import HttpResponse


def workbook_from_rows(rows: Iterable[dict]):
    workbook = Workbook()
    sheet = workbook.active
    rows = list(rows)
    if not rows:
        return workbook
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, '') for header in headers])
    return workbook


def excel_response(workbook: Workbook, filename: str):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def load_rows_from_upload(file_obj):
    workbook = load_workbook(filename=file_obj)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [header or '' for header in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append({headers[idx]: value for idx, value in enumerate(row)})
    return data_rows
