from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable
from uuid import uuid4

from django.conf import settings
from openpyxl import Workbook, load_workbook


DATA_DIR = Path(settings.BASE_DIR) / "data" / "hrms"
DOCUMENTS_DIR = Path(settings.BASE_DIR) / "candidates"


@dataclass(frozen=True)
class ExcelConfig:
    path: Path
    headers: tuple[str, ...]


CANDIDATE_HEADERS = (
    "candidate_id",
    "created_at",
    "candidate_name",
    "mobile_number",
    "email",
    "position_applied_for",
    "source",
    "interview_scheduled",
    "interview_date",
    "interview_status",
    "selection_status",
    "offer_released",
    "joining_date",
    "final_status",
    "remarks",
)

ONBOARDING_HEADERS = (
    "candidate_id",
    "category",
    "document_status",
    "hr_verified",
    "final_onboarding_status",
    "uploaded_documents",
    "optional_documents",
    "notes",
    "submitted_at",
)

CANDIDATES_FILE = ExcelConfig(DATA_DIR / "candidates.xlsx", CANDIDATE_HEADERS)
ONBOARDING_FILE = ExcelConfig(DATA_DIR / "onboarding.xlsx", ONBOARDING_HEADERS)


def ensure_workbook(config: ExcelConfig) -> None:
    config.path.parent.mkdir(parents=True, exist_ok=True)
    if config.path.exists():
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(config.headers))
    workbook.save(config.path)


def load_rows(config: ExcelConfig) -> list[dict]:
    if not config.path.exists():
        return []
    workbook = load_workbook(config.path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(header) for header in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append({headers[idx]: (value if value is not None else "") for idx, value in enumerate(row)})
    return data_rows


def append_row(config: ExcelConfig, row: dict) -> None:
    ensure_workbook(config)
    workbook = load_workbook(config.path)
    sheet = workbook.active
    sheet.append([row.get(header, "") for header in config.headers])
    workbook.save(config.path)


def create_candidate_id() -> str:
    return f"CAND-{uuid4().hex[:10].upper()}"


def candidate_exists(candidate_id: str) -> dict | None:
    for row in load_rows(CANDIDATES_FILE):
        if row.get("candidate_id") == candidate_id:
            return row
    return None


def save_files(candidate_id: str, category: str, files: Iterable[tuple[str, list]]) -> list[str]:
    stored_paths: list[str] = []
    base_dir = DOCUMENTS_DIR / candidate_id / "documents" / category
    base_dir.mkdir(parents=True, exist_ok=True)
    for field_name, field_files in files:
        if not field_files:
            continue
        field_dir = base_dir / field_name
        field_dir.mkdir(parents=True, exist_ok=True)
        for index, file_obj in enumerate(field_files, start=1):
            safe_name = Path(file_obj.name).name
            target = field_dir / f"{index}_{safe_name}"
            with target.open("wb") as destination:
                for chunk in file_obj.chunks():
                    destination.write(chunk)
            stored_paths.append(str(target))
    return stored_paths


def build_onboarding_row(
    *,
    candidate_id: str,
    category: str,
    document_status: str,
    hr_verified: str,
    final_onboarding_status: str,
    uploaded_documents: str,
    optional_documents: str,
    notes: str,
) -> dict:
    return {
        "candidate_id": candidate_id,
        "category": category,
        "document_status": document_status,
        "hr_verified": hr_verified,
        "final_onboarding_status": final_onboarding_status,
        "uploaded_documents": uploaded_documents,
        "optional_documents": optional_documents,
        "notes": notes,
        "submitted_at": datetime.utcnow().isoformat(),
    }
